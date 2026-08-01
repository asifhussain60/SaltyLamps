#!/usr/bin/env python3
"""Builds the catalogue workbook the shop owner fills in.

WHY THIS EXISTS
---------------
The new Cloudflare site starts from a fresh inventory. Rather than us inferring
prices from a price list dated April 2025 — which already left 13 items unpriceable
and hid two duplicate product codes — the owner confirms every cost, stock level and
code himself, in one document, once.

He enters TRADE COST. The shop price is calculated from it in the sheet, live, using
the same rule the shop already uses (double the cost, rounded up to the next whole
pound, less a penny). That means he sees the customer-facing price as he types, the
margin stays visible, and a future cost change is one column edit rather than a
repricing exercise.

The workbook is deliberately opinionated about clarity: one instruction tab that has
to be read first, three colours that mean exactly one thing each, a worked example
row, and per-column notes. A spreadsheet sent to a non-technical recipient fails on
ambiguity far more often than on missing features.

USAGE
    python3 scripts/build_owner_workbook.py
    (reads the live catalogue from the running dev server)
"""

import json
import re
import urllib.request
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent.parent
API = "http://localhost:8788"
OUT = ROOT / "handover" / f"Salty-Lamps-Catalogue-{date.today().isoformat()}.xlsx"

# ---------------------------------------------------------------------------
# House style. Three fills, three meanings, stated on the instructions tab and
# repeated in every header row. Nothing else is coloured, so colour always means
# something.

FILL_FILL_IN = PatternFill("solid", fgColor="DCE9F7")   # blue  — you fill this in
FILL_READONLY = PatternFill("solid", fgColor="EFEFEF")  # grey  — for reference
FILL_CALC = PatternFill("solid", fgColor="E2F0DA")      # green — calculated
FILL_HEADER = PatternFill("solid", fgColor="2F3A40")
FILL_WARN = PatternFill("solid", fgColor="FCE8E6")
FILL_TITLE = PatternFill("solid", fgColor="D88A35")

HEAD = Font(bold=True, color="FFFFFF", size=11)
TITLE = Font(bold=True, size=16, color="1C110D")
BOLD = Font(bold=True)
MUTED = Font(color="6B6B6B", italic=True)
THIN = Side(style="thin", color="C9C9C9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")


def fetch(path):
    with urllib.request.urlopen(f"{API}{path}", timeout=20) as r:
        return json.load(r)


def pack_size(label):
    """'1-1.5Kg / 12pc' -> 12. A listing sold as a pack of N is priced as N units."""
    m = re.search(r"(\d+)\s*pc\b", label or "", re.I)
    return int(m.group(1)) if m else 1


# Trade costs from the 01 April 2025 list, prepopulated so the owner confirms or
# corrects rather than starting from a blank column. Blank where the list has no
# line — those rows are the ones that most need his attention.
TRADE_2025 = {
    "TC-1": 2.20, "TC-2": 3.00, "TCH": 3.50, "TCA": 3.50, "TCB": 3.50,
    "TCG": 3.50, "TCC": 3.50, "TCCL": 4.00,
    "RSL-XS": 6.00, "RSL-1": 7.50, "RSL-2": 12.50, "RSL-3": 14.00,
    "RSL-B": 12.50, "RSL-BB": 18.00, "RSL-E": 12.50, "RSL-M": 12.50,
    "RSL-BL": 12.50, "RSL-T": 12.50, "RSL-F": 15.00, "RSL-P": 12.50,
    "RSL-D": 15.00, "RSL-A": 15.00, "RSL-B1": 12.50, "RSL-B2": 12.50,
    "RSL-B3": 15.00, "RSL-WB": 18.00,
    "CUL-100": 0.60, "CUL-500": 1.25, "CUL-1000": 1.50, "CUL-5000": 7.00,
    "SB-01": 1.85, "SH-01": 1.85, "SD-01": 1.85, "SB-02": 2.00, "SB-02L": 2.00,
    "SS-01": 2.20, "SS-01-02": 2.50, "SI-01": 6.00,
    "SPC-8": 10.00, "SPS-8": 10.00, "SPR-12": 12.50, "HSP-1": 8.00,
    "SB-03": 8.00, "SB-04": 10.00, "SB-05": 2.50, "SB-06": 12.00,
    "SG-01": 1.80, "ST-841": 3.50, "ST-842": 4.00,
    "E14-1": 0.70, "E14-2": 0.80, "PC": 3.50, "PC-1": 5.00,
    "SL-1": 2.50, "SL-2": 3.50, "SL-3": 5.00,
}

# Same normalisation the repricing script uses, so the workbook and the site can
# never disagree about which trade-list line a product code refers to.
#   'SL-1 (12)' -> 'SL-1'      pack size lives in the variant label, not the code
#   'CUL-1000C' -> 'CUL-1000'  the C is grind (coarse), same cost
def normalise_code(sku_code):
    code = re.sub(r"\s*\(\d+\)\s*$", "", str(sku_code)).strip()
    if re.match(r"^CUL-\d+C$", code):
        code = code[:-1]
    return code


# Per-row notes for the things we already know are wrong, so the owner is asked
# about them on the exact line they concern rather than in a separate list he has
# to cross-reference.
def row_note(sku, product, dup_owner):
    code = normalise_code(sku["sku"])
    if re.match(r"^product_[0-9a-f-]+-default$", sku["sku"]):
        return ("NEEDS A REAL CODE. This is a leftover from the old website's export, "
                "not a product code. Please give it a proper one.")
    if sku["sku"] in ("Deal 1", "Deal 2"):
        return ("Bundle offer. Please enter what the bundle costs you in total, and "
                "tell us in the Notes what's included.")
    # Only the product that ISN'T the code's rightful owner gets asked. Flagging
    # both sides would put a scary note on rows that are perfectly fine.
    if code in dup_owner and dup_owner[code] != product["name"]:
        return (f"NEEDS ITS OWN CODE. '{code}' is already the code for "
                f"\"{dup_owner[code]}\", so the two can't be told apart. "
                "Please give this one a different code.")
    if code not in TRADE_2025:
        return "No cost on the 2025 list — please add one."
    if "Watt" in (sku["variant_label"] or ""):
        watt = re.search(r"(\d+)\s*Watt", sku["variant_label"]).group(1)
        expected = "E14-1" if watt == "15" else "E14-2"
        if code != expected:
            return (f"CODE LOOKS WRONG. This is the {watt}W bulb but it's coded {code}; "
                    f"the list says {watt}W is {expected}. Please confirm.")
    return ""


def style_header(ws, headers, row=1):
    for i, (label, width, fill) in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=label)
        c.font = HEAD
        c.fill = FILL_HEADER
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        c.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[row].height = 34


def build():
    products = fetch("/api/admin/products")["products"]
    categories = fetch("/api/admin/categories")["categories"]

    # The owner should never see an internal slug. Map them to the names he'd
    # recognise from the shop, and drop the catch-all bucket, which means nothing
    # to him.
    cat_name = {c["slug"]: c["name"] for c in categories}

    def friendly_categories(csv):
        names = [cat_name.get(s, s) for s in (csv or "").split(",")
                 if s and s != "all-products"]
        return ", ".join(names) if names else "—"

    # Codes used by two unrelated products. The trade list makes the rightful owner
    # unambiguous in both cases, so only the other product is asked to change.
    RIGHTFUL_OWNER = {
        "SL-2": "Himalayan Rock Salt Lick for Equestrian & Cattle",
        "ST-841": "Himalayan Rock Salt Bricks for Salt Walls",
    }
    per_code = {}
    for p in products:
        for s in p["skus"]:
            per_code.setdefault(normalise_code(s["sku"]), set()).add(p["name"])
    dup_owner = {c: RIGHTFUL_OWNER.get(c, sorted(names)[0])
                 for c, names in per_code.items() if len(names) > 1}

    wb = Workbook()

    # ---------------------------------------------------------------- START HERE
    ws = wb.active
    ws.title = "START HERE"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 108

    def line(text, font=None, fill=None, height=None):
        r = ws.max_row + 1
        c = ws.cell(row=r, column=2, value=text)
        c.alignment = WRAP
        if font:
            c.font = font
        if fill:
            c.fill = fill
        if height:
            ws.row_dimensions[r].height = height
        return r

    ws["B1"] = "Salty Lamps — product and price update"
    ws["B1"].font = TITLE
    ws["B1"].fill = FILL_TITLE
    ws.row_dimensions[1].height = 34
    line("")
    line("We're rebuilding the Salty Lamps website and starting the stock records fresh. "
         "This sheet is how we get the correct, current numbers straight from you rather "
         "than guessing from the April 2025 price list.", height=32)
    line("")
    line("It should take about 20–30 minutes. Nothing here is technical.", font=BOLD)
    line("")

    line("WHAT TO DO", font=Font(bold=True, size=13))
    line("1.  Open the 'Products' tab. There is one row for every item and size you sell — 77 rows.", height=18)
    line("2.  For each row, fill in the BLUE columns:", height=18)
    line("        •  Your cost per unit  —  what you pay for one, before VAT", height=18)
    line("        •  Stock you have now  —  how many you actually hold today", height=18)
    line("        •  Still selling?  —  choose Yes or No from the dropdown", height=18)
    line("        •  Product code  —  only if the one shown is wrong", height=18)
    line("        •  Or set your own shop price  —  only if you don't want the calculated one", height=18)
    line("3.  Costs are already filled in from the April 2025 list. Please correct any that "
         "have changed, and fill in the ones that are blank.", height=30)
    line("4.  Read the 'Notes' column on the right. About 15 rows have a specific question "
         "for you — mostly product codes that are duplicated or missing.", height=30)
    line("5.  Fill in the 'VAT & Delivery' tab. It's 9 questions and it's the part we "
         "genuinely cannot answer for you.", height=30)
    line("6.  Glance at the 'Categories' tab and tell us if any names are wrong.", height=18)
    line("7.  Send the file back. That's it.", font=BOLD, height=18)
    line("")

    line("THE COLOURS MEAN SOMETHING", font=Font(bold=True, size=13))
    r = line("     Blue  =  please fill this in.  These are the only cells you need to touch.")
    ws.cell(row=r, column=2).fill = FILL_FILL_IN
    r = line("     Green =  worked out automatically.  Please don't type in these — they update themselves.")
    ws.cell(row=r, column=2).fill = FILL_CALC
    r = line("     Grey  =  for your reference only.  This is what the website currently shows.")
    ws.cell(row=r, column=2).fill = FILL_READONLY
    line("")
    r = line("Please don't delete or edit column A ('Ref'). It's how we match your answers back "
             "to the right product — you can sort and filter freely, but if that column goes, "
             "we lose the link.", height=30)
    ws.cell(row=r, column=2).font = Font(bold=True, color="9B4328")
    line("")

    line("HOW THE SHOP PRICE IS WORKED OUT", font=Font(bold=True, size=13))
    line("You enter the cost; the green 'New shop price' column fills itself in as you type. "
         "It doubles your cost, then rounds to the nearest .99 — which is exactly what your "
         "current prices already do. A £15 lamp cost becomes £29.99, a £12.50 cost becomes "
         "£24.99. If a pack contains several units, it multiplies up first.", height=46)
    line("")
    r = line("If you'd rather set a different shop price for something, put it in the "
             "'Or set your own shop price' column and we'll use your figure instead of the "
             "calculated one.", height=30)
    ws.cell(row=r, column=2).font = MUTED
    line("")

    line("WHY WE'RE ASKING", font=Font(bold=True, size=13))
    r = line("When the old website's products were exported, the per-size prices were lost. "
             "Every size of an item ended up with the same price — a twelve-pack of large salt "
             "licks was priced the same as a single small one, and ten kilos of culinary salt "
             "the same as one kilo. Twenty-one items were being sold for less than they cost you. "
             "We've corrected those from the 2025 list, but we'd rather rebuild from your real "
             "figures than from a list that's a year old.", height=76)
    ws.cell(row=r, column=2).fill = FILL_WARN
    line("")
    line(f"Prepared {date.today().strftime('%-d %B %Y')} · 35 products · 77 sizes · "
         "any questions, just ask.", font=MUTED)

    # ------------------------------------------------------------------ PRODUCTS
    ws = wb.create_sheet("Products")
    ws.freeze_panes = "A3"
    headers = [
        ("Ref\n(don't change)", 7, FILL_READONLY),
        ("Product", 42, None),
        ("Size / option", 18, None),
        ("Product code\n(correct it if wrong)", 16, FILL_FILL_IN),
        ("Category", 20, None),
        ("Price on the site now\n(reference only)", 14, FILL_READONLY),
        ("YOUR COST per unit £\n(before VAT)", 15, FILL_FILL_IN),
        ("Units in\nthis pack", 9, None),
        ("Shop price\n(worked out for you)", 15, FILL_CALC),
        ("Or set your own\nshop price £", 14, FILL_FILL_IN),
        ("STOCK you\nhave now", 11, FILL_FILL_IN),
        ("STILL\nSELLING?", 11, FILL_FILL_IN),
        ("Notes — please read, some rows have a question for you", 58, None),
    ]

    ws["A1"] = ("Fill in the BLUE columns. The green one fills itself in as you type. Grey is reference only — "
                "please leave column A alone, it's how we match your answers back up.")
    ws["A1"].font = Font(bold=True, size=11, color="9B4328")
    ws.merge_cells("A1:M1")
    ws.row_dimensions[1].height = 22

    style_header(ws, headers, row=2)

    yes_no = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    yes_no.error = "Please choose Yes or No."
    yes_no.prompt = "Are you still selling this size?"
    ws.add_data_validation(yes_no)

    row = 3
    first_data_row = row
    for product in sorted(products, key=lambda p: p["name"]):
        for sku in product["skus"]:
            code = normalise_code(sku["sku"])
            qty = pack_size(sku["variant_label"])
            note = row_note(sku, product, dup_owner)

            # Column A is the join key. Codes can be duplicated or corrected by the
            # owner, and rows can be sorted or filtered, so matching on either
            # would silently attach one product's cost to another.
            ref = ws.cell(row=row, column=1, value=sku["id"])
            ref.fill = FILL_READONLY
            ref.font = Font(color="9B9B9B", size=9)
            ref.alignment = Alignment(horizontal="center")

            ws.cell(row=row, column=2, value=product["name"]).alignment = WRAP
            ws.cell(row=row, column=3, value=sku["variant_label"] or "—")

            skucode = ws.cell(row=row, column=4, value=sku["sku"])
            skucode.fill = FILL_FILL_IN

            ws.cell(row=row, column=5, value=friendly_categories(product["categories"])).alignment = WRAP

            cur = ws.cell(row=row, column=6, value=sku["price_pence"] / 100)
            cur.number_format = '£#,##0.00'
            cur.fill = FILL_READONLY

            cost = ws.cell(row=row, column=7, value=TRADE_2025.get(code))
            cost.number_format = '£#,##0.00'
            cost.fill = FILL_FILL_IN

            ws.cell(row=row, column=8, value=qty).alignment = Alignment(horizontal="center")

            # Live formula, so he sees the customer-facing price as he types the cost.
            calc = ws.cell(row=row, column=9,
                           value=f'=IF(G{row}="","",CEILING(G{row}*H{row}*2,1)-0.01)')
            calc.number_format = '£#,##0.00'
            calc.fill = FILL_CALC

            override = ws.cell(row=row, column=10)
            override.number_format = '£#,##0.00'
            override.fill = FILL_FILL_IN

            stock = ws.cell(row=row, column=11)
            stock.fill = FILL_FILL_IN
            selling = ws.cell(row=row, column=12)
            selling.fill = FILL_FILL_IN
            yes_no.add(selling)

            n = ws.cell(row=row, column=13, value=note)
            n.alignment = WRAP
            if note:
                n.fill = FILL_WARN
                n.font = Font(bold=True, color="9B4328")

            for col in range(1, 14):
                ws.cell(row=row, column=col).border = BORDER
            row += 1

    ws.auto_filter.ref = f"A2:M{row - 1}"

    # ------------------------------------------------------------- VAT & DELIVERY
    ws = wb.create_sheet("VAT & Delivery")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 62
    ws.column_dimensions["C"].width = 26
    ws.column_dimensions["D"].width = 58

    ws["B1"] = "VAT and delivery"
    ws["B1"].font = TITLE
    ws["B1"].fill = FILL_TITLE
    ws.merge_cells("B1:D1")
    ws.row_dimensions[1].height = 30

    ws["B3"] = ("The website currently charges nothing for delivery and shows no VAT anywhere. "
                "Neither can stay that way once it takes real orders, and these are the questions "
                "only you can answer.")
    ws["B3"].alignment = WRAP
    ws.merge_cells("B3:D3")
    ws.row_dimensions[3].height = 34

    for col, label in (("B", "Question"), ("C", "YOUR ANSWER"), ("D", "Why we're asking")):
        c = ws[f"{col}5"]
        c.value = label
        c.font = HEAD
        c.fill = FILL_HEADER
        c.alignment = Alignment(horizontal="center", vertical="center")

    questions = [
        ("Are you VAT registered?", "Yes / No",
         "Decides whether VAT has to appear anywhere on the site at all."),
        ("If yes, your VAT number", "",
         "It has to be shown on the website and on receipts."),
        ("Should the prices customers see include VAT?", "Yes / No",
         "UK rules say prices shown to the public must include VAT. Answer No only if you sell mainly to trade."),
        ("Standard UK delivery charge", "£",
         "Charged on a normal order. Right now every order ships free."),
        ("Free delivery over this amount?", "£  (or 'none')",
         "A common way to lift order sizes. Leave as 'none' if you'd rather not."),
        ("Delivery charge for heavy items\n(bricks, 12-packs, 10kg salt)", "£",
         "A 12-pack of 4-5kg licks is around 50kg. Standard postage won't cover it."),
        ("Do you deliver outside the UK?", "Yes / No",
         "The checkout currently only accepts UK addresses."),
        ("If yes, which countries and what charge?", "",
         "We'll add them to the checkout."),
        ("Warn you when stock drops to how many?", "  (currently 5)",
         "The admin screen highlights anything at or below this number."),
    ]
    r = 6
    for q, placeholder, why in questions:
        qc = ws.cell(row=r, column=2, value=q)
        qc.alignment = WRAP
        qc.font = BOLD
        ac = ws.cell(row=r, column=3, value=placeholder)
        ac.fill = FILL_FILL_IN
        ac.alignment = Alignment(horizontal="center", vertical="center")
        ac.font = MUTED
        wc = ws.cell(row=r, column=4, value=why)
        wc.alignment = WRAP
        for col in (2, 3, 4):
            ws.cell(row=r, column=col).border = BORDER
        ws.row_dimensions[r].height = 40
        r += 1

    # ---------------------------------------------------------------- CATEGORIES
    ws = wb.create_sheet("Categories")
    ws.freeze_panes = "A3"
    ws["A1"] = "These are the sections the shop is organised into. Tell us if any name is wrong or a section should go."
    ws["A1"].font = Font(bold=True, size=11, color="9B4328")
    ws.merge_cells("A1:E1")
    style_header(ws, [
        ("Section name", 34, None),
        ("Products in it", 14, None),
        ("Shown on the site?", 16, None),
        ("KEEP IT?", 12, FILL_FILL_IN),
        ("Your notes", 56, FILL_FILL_IN),
    ], row=2)

    keep_yn = DataValidation(type="list", formula1='"Yes,No,Rename"', allow_blank=True)
    ws.add_data_validation(keep_yn)

    r = 3
    for cat in categories:
        if cat["is_virtual"]:
            continue
        ws.cell(row=r, column=1, value=cat["name"])
        ws.cell(row=r, column=2, value=cat["product_count"]).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=3, value="Yes" if cat["visible"] else "No — currently hidden")
        k = ws.cell(row=r, column=4)
        k.fill = FILL_FILL_IN
        keep_yn.add(k)
        note = ws.cell(row=r, column=5)
        note.fill = FILL_FILL_IN
        if cat["product_count"] == 0:
            note.value = "Nothing is in this section yet — is it still coming?"
            note.font = MUTED
        for col in range(1, 6):
            ws.cell(row=r, column=col).border = BORDER
        r += 1

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT.relative_to(ROOT)}")
    print(f"  Products tab: {row - 3} product lines across {len(products)} products")
    print(f"  Rows carrying a question: {sum(1 for p in products for s in p['skus'] if row_note(s, p, dup_owner))}")


if __name__ == "__main__":
    build()
