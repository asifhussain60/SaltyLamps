#!/usr/bin/env python3
"""Reads the completed catalogue workbook back in.

The other half of build_owner_workbook.py. The owner fills in cost, stock and
whether he's still selling each line; this validates what came back and writes it
through the admin API so every change lands in the audit log.

DESIGN NOTES
------------
Matched on the hidden 'Ref' column (column A), not on row position and not on
product code. The owner may sort or filter rows, and two products currently share a
code — matching on either would silently attach one product's cost to another,
which is the worst possible failure for a pricing import. Ref is the database's own
identifier and never changes.

Shop price is recalculated here rather than read from the sheet's green column.
openpyxl cannot evaluate formulas, so that column is unreadable unless Excel has
saved cached values, and trusting a cached value would mean a stale figure could
slip in. Recomputing from cost applies the same rule the sheet displays.

Reports first, writes only with --apply. Nothing is ever deleted: a line marked
"No" is hidden, not removed, because order history references it.

TARGETING A SECOND DATABASE
---------------------------
skus.id is AUTOINCREMENT, so the Ref numbers mean nothing in any database other
than the one the workbook was built from. The test site's ids happen to sit 77
higher than the dev shop's, and matching Ref straight against it would attach one
product's price to another. --ref-source therefore takes a snapshot of the shop the
Refs came from and resolves each Ref through the pair (product id, code, size),
which does travel. The snapshot must be taken BEFORE any import runs, because the
import changes seven of the codes it matches on.

USAGE
    python3 scripts/import_owner_workbook.py <file.xlsx>
    python3 scripts/import_owner_workbook.py <file.xlsx> --apply

    --api=<url>            shop to write to        (default http://localhost:8788)
    --ref-source=<path>    JSON snapshot of the shop the Ref column came from,
                           as returned by GET /api/admin/products
    --set-code=REF:CODE    override one row's product code (repeatable). For codes
                           the owner was asked to fix and did not — the duplicate
                           check below refuses to write while any remain.
"""

import json
import math
import re
import sys
import urllib.error
import urllib.request
from pathlib import Path

from openpyxl import load_workbook

API = "http://localhost:8788"

# Cloudflare answers 403 to urllib's default "Python-urllib/3.x", so every request
# has to name itself. Without this the deployed site simply refuses to talk.
HEADERS = {"user-agent": "salty-lamps-workbook-importer/1.0"}


def fetch(path):
    req = urllib.request.Request(f"{API}{path}", headers=HEADERS)
    with urllib.request.urlopen(req, timeout=30) as r:
        return json.load(r)


def patch(path, body):
    req = urllib.request.Request(
        f"{API}{path}", method="PATCH",
        data=json.dumps(body).encode(),
        headers={**HEADERS, "content-type": "application/json"},
    )
    try:
        with urllib.request.urlopen(req, timeout=30) as r:
            return True, json.load(r)
    except urllib.error.HTTPError as e:
        return False, e.read().decode()[:200]


def shop_pence(cost, pack):
    """The rule the sheet shows live: double the cost, up to the next pound, less 1p."""
    return math.ceil(cost * pack * 2) * 100 - 1


def variant_key(product, sku):
    """The identity of a variant that survives being copied into another database.

    Not skus.id, which is AUTOINCREMENT and local. products.id is stable TEXT, and
    the code and the size together separate the variants within a product — the same
    pair catalogue-reset.mjs keys on, for the same reason.
    """
    return (
        product["id"],
        str(sku.get("sku") or "").strip(),
        str(sku.get("variant_label") or "").strip(),
    )


def resolve_refs(target, ref_source):
    """Map every Ref in the sheet to the row it means in the shop being written.

    With no snapshot the Refs are this shop's own ids and stand for themselves. With
    one, they name rows in a different database and are resolved through variant_key.
    A product holding a single variant falls back to matching on the product alone,
    which is what rescues the one line already renamed by hand on the test site.
    """
    if not ref_source:
        return {s["id"]: (p, s) for p in target for s in p["skus"]}, []

    source = json.loads(Path(ref_source).read_text())["products"]
    by_key = {variant_key(p, s): (p, s) for p in target for s in p["skus"]}
    only_child = {p["id"]: (p, p["skus"][0]) for p in target if len(p["skus"]) == 1}

    resolved, problems = {}, []
    for p in source:
        for s in p["skus"]:
            hit = by_key.get(variant_key(p, s))
            if hit is None and len(p["skus"]) == 1:
                hit = only_child.get(p["id"])
            if hit is None:
                problems.append(f"Ref {s['id']} ({s['sku']} — {p['name']}) has no matching "
                                f"line at {API}")
            else:
                resolved[s["id"]] = hit
    return resolved, problems


def main():
    global API
    argv = sys.argv[1:]
    positional = [a for a in argv if not a.startswith("--")]
    if not positional:
        sys.exit("usage: import_owner_workbook.py <file.xlsx> [--apply] [--api=<url>] "
                 "[--ref-source=<path>] [--set-code=REF:CODE]")
    path = Path(positional[0])
    apply = "--apply" in argv

    def flag(name):
        hit = next((a for a in argv if a.startswith(f"--{name}=")), None)
        return hit.split("=", 1)[1] if hit else None

    API = (flag("api") or API).rstrip("/")
    ref_source = flag("ref-source")

    # Codes the owner was asked to fix and did not. Supplied per row rather than
    # hardcoded, so the script never carries a guess about anyone's product.
    forced_codes = {}
    for a in argv:
        if a.startswith("--set-code="):
            ref, _, code = a.split("=", 1)[1].partition(":")
            try:
                forced_codes[int(ref)] = code.strip()
            except ValueError:
                sys.exit(f"--set-code needs REF:CODE, got {a.split('=', 1)[1]!r}")

    if not path.exists():
        sys.exit(f"No such file: {path}")
    if ref_source and not Path(ref_source).exists():
        sys.exit(f"No such snapshot: {ref_source}")

    ws = load_workbook(path, data_only=False)["Products"]

    # Row 1 is the reminder banner, row 2 the header, data from row 3.
    sheet = {}
    problems = []
    for r in range(3, ws.max_row + 1):
        ref = ws.cell(row=r, column=1).value
        if ref in (None, ""):
            continue
        try:
            ref = int(ref)
        except (TypeError, ValueError):
            problems.append(f"row {r}: the Ref in column A is {ref!r}, which isn't one of ours")
            continue
        if ref in sheet:
            problems.append(f"row {r}: Ref {ref} appears more than once — the sheet has been "
                            "copied or duplicated somewhere")
            continue
        sheet[ref] = {
            "row": r,
            "product": ws.cell(row=r, column=2).value,
            "variant": ws.cell(row=r, column=3).value,
            "code": ws.cell(row=r, column=4).value,
            "cost": ws.cell(row=r, column=7).value,
            "pack": ws.cell(row=r, column=8).value or 1,
            "override": ws.cell(row=r, column=10).value,
            "stock": ws.cell(row=r, column=11).value,
            "selling": str(ws.cell(row=r, column=12).value or "").strip().lower(),
            "note": ws.cell(row=r, column=13).value,
        }

    live = fetch("/api/admin/products")["products"]
    by_ref, map_problems = resolve_refs(live, ref_source)
    problems.extend(map_problems)

    price_changes, stock_changes, code_changes, hide, missing, blank = [], [], [], [], [], []
    size_changes, mode_switches = [], []

    # A single row can change its price, its code, its size and its stock at once.
    # The variant endpoint replaces the whole record, so one PATCH per field would
    # quietly undo the fields written before it — the second call still carries the
    # first call's stale values. Everything a row wants is merged here and sent once.
    wanted = {}

    def want(sku, field, value):
        wanted.setdefault(sku["id"], {})[field] = value

    for ref, row in sheet.items():
        if ref not in by_ref:
            missing.append((ref, row))
            continue
        product, sku = by_ref[ref]
        code = forced_codes.get(ref) or (str(row["code"]).strip() if row["code"] else None)
        label = str(code or sku["sku"])

        # An explicit shop price always wins over the calculated one — that column
        # exists precisely so the owner can disagree with the formula.
        if row["override"] not in (None, ""):
            try:
                want_pence = round(float(row["override"]) * 100)
            except (TypeError, ValueError):
                problems.append(f"row {row['row']}: your own shop price {row['override']!r} isn't a number")
                continue
            if want_pence != sku["price_pence"]:
                want(sku, "price_pence", want_pence)
                price_changes.append((label, sku, row, None, want_pence))
        elif row["cost"] in (None, ""):
            blank.append((label, row))
        else:
            try:
                cost = float(row["cost"])
            except (TypeError, ValueError):
                problems.append(f"row {row['row']}: cost {row['cost']!r} is not a number")
                continue
            if cost <= 0:
                problems.append(f"row {row['row']}: cost must be more than zero")
                continue
            want_pence = shop_pence(cost, int(row["pack"]))
            if want_pence != sku["price_pence"]:
                want(sku, "price_pence", want_pence)
                price_changes.append((label, sku, row, cost, want_pence))

        if code and code != sku["sku"]:
            want(sku, "sku", code)
            code_changes.append((sku, sku["sku"], code, ref in forced_codes))

        # The size column started as an em-dash for anything with no size. Where the
        # owner replaced it with a real one, take it. Where he echoed the product code
        # back into it, don't — that would put a code where a customer reads a size.
        size = str(row["variant"] or "").strip()
        if size and size not in ("—", "-") and size != str(row["code"] or "").strip():
            if size != sku["variant_label"]:
                want(sku, "variant_label", size)
                size_changes.append((label, sku["variant_label"], size))

        if row["stock"] not in (None, ""):
            try:
                qty = int(row["stock"])
            except (TypeError, ValueError):
                problems.append(f"row {row['row']}: stock {row['stock']!r} is not a whole number")
            else:
                # A count is a count. A line tracked as a plain yes/no throws the
                # number away on save, so counting it means switching how it's tracked.
                if sku["track_mode"] != "quantity":
                    want(sku, "track_mode", "quantity")
                    mode_switches.append((label, qty))
                if qty != sku["quantity"] or sku["track_mode"] != "quantity":
                    want(sku, "quantity", qty)
                    want(sku, "in_stock", 1 if qty > 0 else 0)
                    stock_changes.append((label, sku, qty))

        if row["selling"] == "no":
            hide.append((label, product, sku))

    for ref in by_ref:
        if ref not in sheet:
            p, s = by_ref[ref]
            problems.append(f"{s['sku']!r} ({p['name']}) is in the shop but its row is missing "
                            "from the sheet")

    # Two products sharing a code is the defect this whole exercise exists to fix, so
    # do not let the fix reintroduce it. Codes may legitimately repeat WITHIN one
    # product (pack sizes of the same item), so this compares across products only.
    code_to_products = {}
    for ref, row in sheet.items():
        effective = forced_codes.get(ref) or (str(row["code"]).strip() if row["code"] else None)
        if ref not in by_ref or not effective:
            continue
        base = re.sub(r"\s*\(\d+\)\s*$", "", effective)
        code_to_products.setdefault(base, set()).add(by_ref[ref][0]["name"])
    for base, names in sorted(code_to_products.items()):
        if len(names) > 1:
            problems.append(f"code {base!r} is used by more than one product: "
                            + "; ".join(sorted(names)))

    g = lambda p: f"£{p / 100:.2f}"
    print(f"\n=== {path.name} → {API} ===\n")
    if ref_source:
        print(f"  refs resolved through  : {ref_source}")
    print(f"  rows read              : {len(sheet)}")
    print(f"  lines to write         : {len(wanted)}")
    print(f"  price changes          : {len(price_changes)}")
    print(f"  product code changes   : {len(code_changes)}")
    print(f"  size name changes      : {len(size_changes)}")
    print(f"  stock changes          : {len(stock_changes)}")
    print(f"  lines marked not selling: {len(hide)}")
    print(f"  costs left blank       : {len(blank)}")
    print(f"  codes not in the shop  : {len(missing)}")

    if problems:
        print(f"\n  !! {len(problems)} problem(s) — nothing will be written until these are resolved:")
        for p in problems[:25]:
            print(f"     {p}")

    if price_changes:
        print("\n--- Price changes ---")
        for code, sku, row, cost, pence in price_changes:
            how = f"cost £{cost:>7.2f} x{row['pack']:<3}" if cost is not None else "price set by hand  "
            print(f"  {code:<16} {how} → {g(pence):>9}  (was {g(sku['price_pence'])})")

    if code_changes:
        print("\n--- Product code changes ---")
        for sku, was, now, forced in code_changes:
            flag_text = "   << set here, NOT by the owner — needs confirming" if forced else ""
            print(f"  {was:<40} → {now}{flag_text}")

    if size_changes:
        print("\n--- Size name changes ---")
        for code, was, now in size_changes:
            print(f"  {code:<16} {was or '(blank)':<28} → {now}")

    if mode_switches:
        print("\n--- Lines becoming counted stock (were a plain yes/no) ---")
        for code, qty in mode_switches:
            print(f"  {code:<16} now counted, starting at {qty}")

    if stock_changes:
        print("\n--- Stock changes ---")
        for code, sku, qty in stock_changes:
            print(f"  {code:<16} {sku['quantity']} → {qty}")

    if hide:
        print("\n--- Marked 'No' — will be HIDDEN, not deleted (order history references them) ---")
        for code, product, sku in hide:
            print(f"  {code:<16} {product['name']}")

    if blank:
        print(f"\n--- {len(blank)} line(s) with no cost given — left exactly as they are ---")
        for code, row in blank[:15]:
            print(f"  {code:<16} {row['product']}")

    if missing:
        print(f"\n--- {len(missing)} code(s) in the sheet that don't exist in the shop ---")
        print("    (probably renamed codes — these need creating by hand, not guessed at)")
        for code, row in missing[:15]:
            print(f"  {code:<16} {row['product']}")

    if problems:
        print("\nRefusing to write while there are problems above. Fix the sheet and re-run.\n")
        sys.exit(1)

    if not apply:
        print("\nReport only. Re-run with --apply to write these through the admin API.\n")
        return

    # One call per line, carrying everything that line changes. See `wanted` above for
    # why this is not one call per field.
    live_skus = {s["id"]: s for _, s in by_ref.values()}
    ok = 0
    for sku_id, fields in wanted.items():
        sku = live_skus[sku_id]
        good, err = patch(f"/api/admin/skus/{sku_id}", {**sku, **fields})
        ok += good
        if not good:
            print(f"  FAILED {sku['sku']}: {err}")
    print(f"\n  {ok} of {len(wanted)} line(s) written.")
    if hide:
        print(f"  {len(hide)} line(s) marked not-selling were NOT changed — hiding a product is "
              "a bigger decision than a price, so do it in the admin.\n")


if __name__ == "__main__":
    main()
